#!/usr/bin/env python3
"""SLT Yield 分析工具：读取 slt_yield 格式 Excel，按工序合并并可视化月度良率。"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import json

import pandas as pd
import matplotlib

matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

plt.rcParams["font.sans-serif"] = ["PingFang SC", "Heiti SC", "SimHei", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False

REQUIRED_COLUMNS = ["工序", "PassQty", "TestQty", "FailQty", "YearMonth"]
BIN_REQUIRED_COLUMNS = ["TEST_STAGE", "BIN_NAME", "YEAR", "MONTH", "FAIL_QTY", "TEST_QTY"]
PASS_BIN_NAMES = {"PASS", "Pass", "pass"}
IGNORE_BIN_NAMES = {"ALL Vector"}
TEMP_LABELS = ("常温", "低温", "高温")
CATEGORY_COLORS = {"常温": "#4472C4", "低温": "#70AD47", "高温": "#ED7D31"}
CONFIG_PATH = Path(__file__).parent / "process_categories.json"
DEFAULT_CATEGORY_MAP = {
    "MT1": "常温",
    "MT2": "常温",
    "MT3": "常温",
    "MT9": "低温",
    "MT10": "高温",
    "MT11": "高温",
}


def load_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {"last_input": "", "files": {}}


def save_config(config: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def file_config_key(filepath: str) -> str:
    return str(Path(filepath).resolve())


def get_saved_categories(filepath: str, config: dict) -> dict[str, str]:
    return config.get("files", {}).get(file_config_key(filepath), {})


def save_categories_for_file(filepath: str, mapping: dict[str, str], config: dict) -> None:
    key = file_config_key(filepath)
    config.setdefault("files", {})[key] = mapping
    config["last_input"] = key
    save_config(config)


def sort_processes(processes: list[str]) -> list[str]:
    def key(name: str):
        digits = "".join(c for c in name if c.isdigit())
        return (name[:2] if len(name) >= 2 else name, int(digits) if digits else 0, name)

    return sorted(processes, key=key)


def detect_file_format(columns: list[str]) -> str:
    cols = set(columns)
    if set(REQUIRED_COLUMNS).issubset(cols):
        return "yield"
    if set(BIN_REQUIRED_COLUMNS).issubset(cols):
        return "bin"
    raise ValueError(
        "无法识别文件格式，需包含 yield 列"
        f"({', '.join(REQUIRED_COLUMNS)}) 或 bin 列"
        f"({', '.join(BIN_REQUIRED_COLUMNS)})"
    )


def read_excel_file(filepath: str) -> tuple[str, pd.DataFrame]:
    raw = pd.read_excel(filepath, engine="openpyxl")
    fmt = detect_file_format(raw.columns.tolist())
    if fmt == "yield":
        return fmt, read_yield_file(filepath, raw)
    return fmt, read_bin_file(filepath, raw)


def read_yield_file(filepath: str, raw: pd.DataFrame | None = None) -> pd.DataFrame:
    df = raw if raw is not None else pd.read_excel(filepath, engine="openpyxl")
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必要列: {', '.join(missing)}")

    df = df.dropna(subset=["工序", "YearMonth"]).copy()
    for col in ("PassQty", "TestQty", "FailQty"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["PassQty", "TestQty"])
    df["PassQty"] = df["PassQty"].astype(int)
    df["TestQty"] = df["TestQty"].astype(int)
    if df["FailQty"].isna().all():
        df["FailQty"] = df["TestQty"] - df["PassQty"]
    else:
        df["FailQty"] = df["FailQty"].fillna(df["TestQty"] - df["PassQty"]).astype(int)

    df["YearMonth"] = df["YearMonth"].astype(str)
    return df


def read_bin_file(filepath: str, raw: pd.DataFrame | None = None) -> pd.DataFrame:
    df = raw if raw is not None else pd.read_excel(filepath, engine="openpyxl")
    missing = [c for c in BIN_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必要列: {', '.join(missing)}")

    df = df.dropna(subset=["TEST_STAGE", "BIN_NAME", "YEAR", "MONTH"]).copy()
    df["TEST_STAGE"] = df["TEST_STAGE"].astype(str)
    df["BIN_NAME"] = df["BIN_NAME"].astype(str)
    for col in ("FAIL_QTY", "TEST_QTY", "YEAR", "MONTH"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["FAIL_QTY", "TEST_QTY", "YEAR", "MONTH"])
    df["FAIL_QTY"] = df["FAIL_QTY"].astype(int)
    df["TEST_QTY"] = df["TEST_QTY"].astype(int)
    df["YEAR"] = df["YEAR"].astype(int)
    df["MONTH"] = df["MONTH"].astype(int)
    df["YearMonth"] = df["YEAR"].astype(str) + "-" + df["MONTH"].astype(str).str.zfill(2)
    df = df[~df["BIN_NAME"].apply(is_ignored_bin)].copy()
    return df


def is_ignored_bin(name: str) -> bool:
    return name.strip() in IGNORE_BIN_NAMES


def is_pass_bin(name: str) -> bool:
    return name.upper() == "PASS" or name in PASS_BIN_NAMES


def _calc_category_month_test_qty(month_data: pd.DataFrame, procs: list[str]) -> int:
    total = 0
    for proc in procs:
        proc_data = month_data[month_data["TEST_STAGE"] == proc]
        if not proc_data.empty:
            total += int(proc_data["TEST_QTY"].iloc[0])
    return total


def prepare_bin_category_chart(
    df: pd.DataFrame, process_category: dict[str, str], category: str
) -> tuple[list[str], list[int], list[str], dict[str, list[float | None]]]:
    """按温度聚合：月度 TestQty 及 Top6 Fail Bin 占比。"""
    procs = sort_processes([p for p, cat in process_category.items() if cat == category])
    sub = df[df["TEST_STAGE"].isin(procs)].copy()
    if sub.empty:
        return [], [], [], {}

    months = sorted(sub["YearMonth"].unique())
    test_qtys = [_calc_category_month_test_qty(sub[sub["YearMonth"] == ym], procs) for ym in months]

    fail_data = sub[~sub["BIN_NAME"].apply(is_pass_bin)]
    top_bins = (
        fail_data.groupby("BIN_NAME")["FAIL_QTY"].sum().sort_values(ascending=False).head(6).index.tolist()
    )

    bin_props: dict[str, list[float | None]] = {b: [] for b in top_bins}
    for ym in months:
        month_data = sub[sub["YearMonth"] == ym]
        total_test = _calc_category_month_test_qty(month_data, procs)
        for b in top_bins:
            if total_test == 0:
                bin_props[b].append(None)
            else:
                fail = int(month_data[month_data["BIN_NAME"] == b]["FAIL_QTY"].sum())
                bin_props[b].append(fail / total_test)

    return months, test_qtys, top_bins, bin_props


def build_bin_test_qty_summary(df: pd.DataFrame, process_category: dict[str, str]) -> pd.DataFrame:
    rows = []
    for cat in TEMP_LABELS:
        months, test_qtys, _, _ = prepare_bin_category_chart(df, process_category, cat)
        for ym, qty in zip(months, test_qtys):
            rows.append({"YearMonth": ym, "温度类型": cat, "TestQty": qty})
    if not rows:
        return pd.DataFrame(columns=["YearMonth", "温度类型", "TestQty"])
    cat_order = {cat: i for i, cat in enumerate(TEMP_LABELS)}
    result = pd.DataFrame(rows)
    result["_sort"] = result["温度类型"].map(cat_order)
    return result.sort_values(["YearMonth", "_sort"]).drop(columns="_sort").reset_index(drop=True)


def build_bin_prop_detail(df: pd.DataFrame, process_category: dict[str, str]) -> pd.DataFrame:
    rows = []
    for cat in TEMP_LABELS:
        months, _, top_bins, bin_props = prepare_bin_category_chart(df, process_category, cat)
        if not months:
            continue
        procs = sort_processes([p for p, c in process_category.items() if c == cat])
        sub = df[df["TEST_STAGE"].isin(procs)]
        for idx, ym in enumerate(months):
            month_data = sub[sub["YearMonth"] == ym]
            total_test = _calc_category_month_test_qty(month_data, procs)
            for bin_name in top_bins:
                prop = bin_props[bin_name][idx]
                fail = int(month_data[month_data["BIN_NAME"] == bin_name]["FAIL_QTY"].sum())
                rows.append(
                    {
                        "YearMonth": ym,
                        "温度类型": cat,
                        "BIN_NAME": bin_name,
                        "FAIL_QTY": fail,
                        "TEST_QTY": total_test,
                        "占比": prop,
                    }
                )
    if not rows:
        return pd.DataFrame(columns=["YearMonth", "温度类型", "BIN_NAME", "FAIL_QTY", "TEST_QTY", "占比"])
    cat_order = {cat: i for i, cat in enumerate(TEMP_LABELS)}
    result = pd.DataFrame(rows)
    result["_sort"] = result["温度类型"].map(cat_order)
    return result.sort_values(["YearMonth", "_sort", "BIN_NAME"]).drop(columns="_sort").reset_index(drop=True)


def build_bin_monthly_wide(df: pd.DataFrame, process_category: dict[str, str], category: str) -> pd.DataFrame:
    months, test_qtys, top_bins, bin_props = prepare_bin_category_chart(df, process_category, category)
    rows = []
    for idx, ym in enumerate(months):
        row: dict = {"YearMonth": ym, "TestQty": test_qtys[idx]}
        for bin_name in top_bins:
            row[bin_name] = bin_props[bin_name][idx]
        rows.append(row)
    return pd.DataFrame(rows)


def _apply_excel_styles(writer, percent_headers: tuple[str, ...] = ("Yield", "总Yield", "占比")):
    from openpyxl.styles import Font

    bold = Font(bold=True)
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.font = bold
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(1, cell.column).value
                if header and any(h in str(header) for h in percent_headers):
                    cell.number_format = "0.00%"
                elif header and any(h in str(header) for h in ("FDPPM", "FAIL_QTY", "TEST_QTY", "TestQty")):
                    cell.number_format = "#,##0"


def export_bin_excel(df: pd.DataFrame, process_category: dict[str, str], output_path: str):
    test_qty = build_bin_test_qty_summary(df, process_category)
    bin_detail = build_bin_prop_detail(df, process_category)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        test_qty.to_excel(writer, sheet_name="TestQty汇总", index=False)
        bin_detail.to_excel(writer, sheet_name="Bin占比明细", index=False)
        for cat in TEMP_LABELS:
            build_bin_monthly_wide(df, process_category, cat).to_excel(
                writer, sheet_name=f"{cat}月度Bin", index=False
            )
        _apply_excel_styles(writer)


def merge_by_process(df: pd.DataFrame, process_category: dict[str, str]) -> pd.DataFrame:
    grouped = (
        df.groupby(["YearMonth", "工序"], as_index=False)
        .agg(PassQty=("PassQty", "sum"), TestQty=("TestQty", "sum"), FailQty=("FailQty", "sum"))
        .sort_values(["YearMonth", "工序"])
    )
    grouped["Yield"] = grouped["PassQty"] / grouped["TestQty"]
    grouped["FDPPM"] = grouped["FailQty"] / grouped["TestQty"] * 1_000_000
    grouped["温度类型"] = grouped["工序"].map(process_category)
    return grouped


def calc_category_yield(month_data: pd.DataFrame, category: str) -> float | None:
    cat_data = month_data[month_data["温度类型"] == category]
    if cat_data.empty or cat_data["TestQty"].sum() == 0:
        return None
    return cat_data["PassQty"].sum() / cat_data["TestQty"].sum()


def calc_total_yield(month_data: pd.DataFrame) -> float | None:
    """总Yield = 常温Yield × 低温Yield × 高温Yield（各温度下 Qty 合并后计算）。"""
    total = 1.0
    for cat in TEMP_LABELS:
        y = calc_category_yield(month_data, cat)
        if y is None:
            return None
        total *= y
    return total


def get_monthly_total_yields(merged: pd.DataFrame) -> tuple[list[str], list[float | None]]:
    months = sorted(merged["YearMonth"].unique())
    totals = [calc_total_yield(merged[merged["YearMonth"] == ym]) for ym in months]
    return months, totals


def get_monthly_total_test_qty(merged: pd.DataFrame, months: list[str]) -> list[int]:
    return [int(merged[merged["YearMonth"] == ym]["TestQty"].sum()) for ym in months]


def get_monthly_test_qty_by_category(merged: pd.DataFrame, months: list[str]) -> dict[str, list[int]]:
    result: dict[str, list[int]] = {cat: [] for cat in TEMP_LABELS}
    for ym in months:
        month_data = merged[merged["YearMonth"] == ym]
        for cat in TEMP_LABELS:
            cat_data = month_data[month_data["温度类型"] == cat]
            result[cat].append(int(cat_data["TestQty"].sum()) if not cat_data.empty else 0)
    return result


def _setup_test_qty_bars(ax, ax_bar, x: list[int], test_qtys: list[int], label: str = "TestQty"):
    """在折线图底层叠加 TestQty 柱状图（右侧 Y 轴）。"""
    ax_bar.bar(x, test_qtys, width=0.55, alpha=0.22, color="#888888", label=label, zorder=1)
    ax_bar.set_ylabel("TestQty")
    ax_bar.yaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
    ax.set_zorder(2)
    ax.patch.set_visible(False)


def _merge_legends(ax, ax_bar, loc="upper right", **kwargs):
    handles1, labels1 = ax.get_legend_handles_labels()
    handles2, labels2 = ax_bar.get_legend_handles_labels()
    ax.legend(handles1 + handles2, labels1 + labels2, **kwargs)


BIN_LINE_COLORS = ["#E74C3C", "#9B59B6", "#F39C12", "#1ABC9C", "#3498DB", "#E67E22"]


def draw_bin_category_chart(
    fig: Figure,
    df: pd.DataFrame,
    process_category: dict[str, str],
    category: str,
):
    fig.clear()
    months, test_qtys, top_bins, bin_props = prepare_bin_category_chart(df, process_category, category)
    if not months:
        ax = fig.add_subplot(111)
        ax.text(0.5, 0.5, f"{category}：无数据", ha="center", va="center", transform=ax.transAxes)
        return

    ax_qty = fig.add_subplot(111)
    x = list(range(len(months)))
    ax_qty.bar(x, test_qtys, width=0.55, alpha=0.35, color=CATEGORY_COLORS[category], label="TestQty", zorder=1)
    ax_qty.set_ylabel("TestQty")
    ax_qty.yaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))

    ax_pct = ax_qty.twinx()
    for i, bin_name in enumerate(top_bins):
        color = BIN_LINE_COLORS[i % len(BIN_LINE_COLORS)]
        props = bin_props[bin_name]
        ax_pct.plot(
            x,
            props,
            marker="o",
            markersize=4,
            label=bin_name,
            color=color,
            linewidth=1.8,
            zorder=3,
        )
        for xi, y in zip(x, props):
            if y is not None:
                ax_pct.annotate(
                    f"{y * 100:.1f}%",
                    (xi, y),
                    textcoords="offset points",
                    xytext=(0, 8 + i * 2),
                    ha="center",
                    fontsize=6,
                    color=color,
                )

    ax_pct.set_ylabel("Fail Bin 占比")
    ax_pct.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter(1.0))
    procs = sort_processes([p for p, c in process_category.items() if c == category])
    ax_qty.set_xticks(x)
    ax_qty.set_xticklabels(months, rotation=45, ha="right")
    ax_qty.set_xlabel("YearMonth")
    ax_qty.set_title(f"{category} TestQty 及 Top{len(top_bins)} Fail Bin 占比")
    _merge_legends(ax_pct, ax_qty, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0, fontsize=7)
    ax_qty.grid(True, alpha=0.3, zorder=0)
    ax_pct.margins(y=0.15)
    fig.subplots_adjust(right=0.72)
    fig.tight_layout(rect=[0, 0, 0.72, 1])


def build_monthly_summary(merged: pd.DataFrame) -> pd.DataFrame:
    months = sorted(merged["YearMonth"].unique())
    processes = sort_processes(merged["工序"].unique().tolist())
    rows = []

    for ym in months:
        month_data = merged[merged["YearMonth"] == ym]
        row = {"YearMonth": ym}
        for proc in processes:
            proc_row = month_data[month_data["工序"] == proc]
            if proc_row.empty:
                row[proc] = None
            else:
                row[proc] = proc_row.iloc[0]["Yield"]
        row["总Yield"] = calc_total_yield(month_data)
        rows.append(row)

    return pd.DataFrame(rows)


def build_category_summary(merged: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for ym in sorted(merged["YearMonth"].unique()):
        month_data = merged[merged["YearMonth"] == ym]
        row = {"YearMonth": ym}
        for cat in TEMP_LABELS:
            cat_data = month_data[month_data["温度类型"] == cat]
            if cat_data.empty:
                row[f"{cat}PassQty"] = 0
                row[f"{cat}TestQty"] = 0
                row[f"{cat}Yield"] = None
            else:
                pass_sum = cat_data["PassQty"].sum()
                test_sum = cat_data["TestQty"].sum()
                row[f"{cat}PassQty"] = pass_sum
                row[f"{cat}TestQty"] = test_sum
                row[f"{cat}Yield"] = pass_sum / test_sum if test_sum else None
        rows.append(row)
    return pd.DataFrame(rows)


def build_temp_qty_merged(merged: pd.DataFrame) -> pd.DataFrame:
    """同温度工序按月份合并 PassQty / TestQty / FailQty。"""
    grouped = (
        merged.groupby(["YearMonth", "温度类型"], as_index=False)
        .agg(PassQty=("PassQty", "sum"), TestQty=("TestQty", "sum"), FailQty=("FailQty", "sum"))
    )
    grouped["Yield"] = grouped["PassQty"] / grouped["TestQty"]
    grouped["FDPPM"] = grouped["FailQty"] / grouped["TestQty"] * 1_000_000
    cat_order = {cat: i for i, cat in enumerate(TEMP_LABELS)}
    grouped["_sort"] = grouped["温度类型"].map(cat_order)
    grouped = grouped.sort_values(["YearMonth", "_sort"]).drop(columns="_sort")
    return grouped[["YearMonth", "温度类型", "PassQty", "TestQty", "FailQty", "Yield", "FDPPM"]]


def export_excel(
    merged: pd.DataFrame,
    monthly: pd.DataFrame,
    category: pd.DataFrame,
    temp_qty: pd.DataFrame,
    output_path: str,
):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="工序汇总", index=False)
        monthly.to_excel(writer, sheet_name="月度良率", index=False)
        category.to_excel(writer, sheet_name="温度分类汇总", index=False)
        temp_qty.to_excel(writer, sheet_name="温度Qty合并", index=False)
        _apply_excel_styles(writer)


class YieldAnalyzerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SLT Yield 分析工具")
        self.geometry("1200x820")
        self.minsize(1000, 700)

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.processes: list[str] = []
        self.process_vars: dict[str, tk.StringVar] = {}
        self.merged_df: pd.DataFrame | None = None
        self.monthly_df: pd.DataFrame | None = None
        self.bin_df: pd.DataFrame | None = None
        self.file_format: str = "yield"
        self.process_category: dict[str, str] = {}
        self.config = load_config()
        self._loading = False
        self.chart_canvases: dict[str, FigureCanvasTkAgg] = {}
        self.chart_figs: dict[str, Figure] = {}

        self._build_ui()
        self._restore_last_session()

    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        file_frame = ttk.LabelFrame(main, text="文件选择", padding=8)
        file_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(file_frame, text="输入文件:").grid(row=0, column=0, sticky=tk.W, padx=(0, 4))
        ttk.Entry(file_frame, textvariable=self.input_path, width=70).grid(row=0, column=1, sticky=tk.EW, padx=4)
        ttk.Button(file_frame, text="浏览...", command=self._browse_input).grid(row=0, column=2, padx=4)

        ttk.Label(file_frame, text="输出文件:").grid(row=1, column=0, sticky=tk.W, padx=(0, 4), pady=(6, 0))
        ttk.Entry(file_frame, textvariable=self.output_path, width=70).grid(
            row=1, column=1, sticky=tk.EW, padx=4, pady=(6, 0)
        )
        ttk.Button(file_frame, text="浏览...", command=self._browse_output).grid(row=1, column=2, padx=4, pady=(6, 0))
        file_frame.columnconfigure(1, weight=1)

        self.process_frame = ttk.LabelFrame(main, text="工序分类（每道工序须且仅能归属一种温度类型）", padding=8)
        self.process_frame.pack(fill=tk.X, pady=(0, 8))

        self.process_inner = ttk.Frame(self.process_frame)
        self.process_inner.pack(fill=tk.X)
        ttk.Label(self.process_inner, text="请先选择输入文件以加载工序列表", foreground="gray").pack(anchor=tk.W)

        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(btn_frame, text="加载工序", command=self._load_processes).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="生成并导出", command=self._generate).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="刷新图表", command=self._refresh_chart).pack(side=tk.LEFT)

        chart_frame = ttk.LabelFrame(main, text="数据可视化", padding=4)
        chart_frame.pack(fill=tk.BOTH, expand=True)

        self.notebook = ttk.Notebook(chart_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self._setup_yield_tabs()

        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(main, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).pack(fill=tk.X, pady=(4, 0))

    def _clear_chart_tabs(self):
        for tab_id in self.notebook.tabs():
            self.notebook.forget(tab_id)
        self.chart_canvases.clear()
        self.chart_figs.clear()

    def _setup_yield_tabs(self):
        self._clear_chart_tabs()
        for key, title in (("all", "全部工序"), ("cat", "温度分类")):
            fig = Figure(figsize=(10, 4), dpi=100)
            canvas = FigureCanvasTkAgg(fig, master=self.notebook)
            self.notebook.add(canvas.get_tk_widget(), text=title)
            self.chart_figs[key] = fig
            self.chart_canvases[key] = canvas

    def _setup_bin_tabs(self):
        self._clear_chart_tabs()
        for cat in TEMP_LABELS:
            fig = Figure(figsize=(10, 4), dpi=100)
            canvas = FigureCanvasTkAgg(fig, master=self.notebook)
            self.notebook.add(canvas.get_tk_widget(), text=f"{cat} Bin")
            self.chart_figs[cat] = fig
            self.chart_canvases[cat] = canvas

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="选择输入 Excel",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )
        if path:
            self.input_path.set(path)
            default_out = Path(path).with_name(Path(path).stem + "_output.xlsx")
            self.output_path.set(str(default_out))
            self._load_processes()

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="选择输出 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            self.output_path.set(path)

    def _restore_last_session(self):
        last = self.config.get("last_input", "")
        if last and Path(last).exists():
            self.input_path.set(last)
            self.output_path.set(str(Path(last).with_name(Path(last).stem + "_output.xlsx")))
            self._load_processes()

    def _on_category_changed(self, *_args):
        if self._loading:
            return
        path = self.input_path.get().strip()
        if not path or not self.process_vars:
            return
        mapping = {proc: self.process_vars[proc].get() for proc in self.processes}
        save_categories_for_file(path, mapping, self.config)

    def _load_processes(self):
        path = self.input_path.get().strip()
        if not path:
            messagebox.showwarning("提示", "请先选择输入文件")
            return
        try:
            fmt, df = read_excel_file(path)
            self.file_format = fmt
            if fmt == "yield":
                self.processes = sort_processes(df["工序"].astype(str).unique().tolist())
                self._setup_yield_tabs()
            else:
                self.processes = sort_processes(df["TEST_STAGE"].astype(str).unique().tolist())
                self._setup_bin_tabs()
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return

        for widget in self.process_inner.winfo_children():
            widget.destroy()

        if not self.processes:
            ttk.Label(self.process_inner, text="未找到有效工序数据", foreground="red").pack(anchor=tk.W)
            return

        header = ttk.Frame(self.process_inner)
        header.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(header, text="工序", width=12, font=("", 10, "bold")).pack(side=tk.LEFT, padx=4)
        for cat in TEMP_LABELS:
            ttk.Label(header, text=cat, width=10, font=("", 10, "bold")).pack(side=tk.LEFT, padx=8)

        self.process_vars.clear()
        saved = get_saved_categories(path, self.config)
        self._loading = True

        for proc in self.processes:
            row = ttk.Frame(self.process_inner)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=proc, width=12).pack(side=tk.LEFT, padx=4)
            if proc in saved and saved[proc] in TEMP_LABELS:
                initial = saved[proc]
            else:
                initial = DEFAULT_CATEGORY_MAP.get(proc, "常温")
            var = tk.StringVar(value=initial)
            var.trace_add("write", self._on_category_changed)
            self.process_vars[proc] = var
            for cat in TEMP_LABELS:
                ttk.Radiobutton(row, text="", variable=var, value=cat).pack(side=tk.LEFT, padx=28)

        self._loading = False
        save_categories_for_file(
            path,
            {proc: self.process_vars[proc].get() for proc in self.processes},
            self.config,
        )

        saved_hint = "（已恢复上次分类）" if saved else ""
        fmt_hint = "Bin格式" if self.file_format == "bin" else "Yield格式"
        self.status_var.set(
            f"[{fmt_hint}] 已加载 {len(self.processes)} 个工序: {', '.join(self.processes)}{saved_hint}"
        )

    def _validate_categories(self) -> dict[str, str] | None:
        if not self.processes:
            messagebox.showwarning("提示", "请先加载工序")
            return None

        mapping = {proc: self.process_vars[proc].get() for proc in self.processes}
        for cat in TEMP_LABELS:
            if not any(v == cat for v in mapping.values()):
                messagebox.showerror("分类错误", f"「{cat}」工序至少需要选择一个")
                return None
        return mapping

    def _generate(self):
        input_path = self.input_path.get().strip()
        output_path = self.output_path.get().strip()
        if not input_path or not output_path:
            messagebox.showwarning("提示", "请选择输入和输出文件路径")
            return

        mapping = self._validate_categories()
        if mapping is None:
            return

        save_categories_for_file(input_path, mapping, self.config)
        self.process_category = mapping

        try:
            if self.file_format == "bin":
                _, df = read_excel_file(input_path)
                self.bin_df = df
                self.merged_df = None
                self.monthly_df = None
                export_bin_excel(df, mapping, output_path)
                self._refresh_chart()
                self.status_var.set(f"已导出: {output_path}  ({len(df)} 行有效 Bin 数据)")
                messagebox.showinfo("完成", f"Bin 分析 Excel 已导出至:\n{output_path}")
                return

            df = read_yield_file(input_path)
            merged = merge_by_process(df, mapping)
            monthly = build_monthly_summary(merged)
            category = build_category_summary(merged)
            temp_qty = build_temp_qty_merged(merged)
            export_excel(merged, monthly, category, temp_qty, output_path)

            self.merged_df = merged
            self.monthly_df = monthly
            self._refresh_chart()
            self.status_var.set(f"已导出: {output_path}  ({len(merged)} 行工序汇总)")
            messagebox.showinfo("完成", f"Excel 已导出至:\n{output_path}")
        except Exception as e:
            messagebox.showerror("处理失败", str(e))

    def _refresh_chart(self):
        mapping = self._validate_categories()
        if mapping is None:
            return
        self.process_category = mapping

        if self.file_format == "bin":
            if self.bin_df is None:
                try:
                    _, self.bin_df = read_excel_file(self.input_path.get().strip())
                except Exception:
                    return
            self._draw_bin_charts()
            return

        if self.merged_df is None or self.monthly_df is None:
            if self.input_path.get().strip() and self.process_vars:
                try:
                    df = read_yield_file(self.input_path.get().strip())
                    self.merged_df = merge_by_process(df, mapping)
                    self.monthly_df = build_monthly_summary(self.merged_df)
                except Exception:
                    return
            else:
                return

        self._draw_process_chart()
        self._draw_category_chart()

    def _draw_bin_charts(self):
        for cat in TEMP_LABELS:
            draw_bin_category_chart(
                self.chart_figs[cat],
                self.bin_df,
                self.process_category,
                cat,
            )
            self.chart_canvases[cat].draw()

    def _draw_process_chart(self):
        self.chart_figs["all"].clear()
        ax = self.chart_figs["all"].add_subplot(111)
        monthly = self.monthly_df
        merged = self.merged_df
        months, total_yields = get_monthly_total_yields(merged)
        x = list(range(len(months)))
        x_labels = months
        test_qtys = get_monthly_total_test_qty(merged, months)

        ax_bar = ax.twinx()
        _setup_test_qty_bars(ax, ax_bar, x, test_qtys)

        for proc in sort_processes(merged["工序"].unique().tolist()):
            cat = merged.loc[merged["工序"] == proc, "温度类型"].iloc[0]
            color = CATEGORY_COLORS.get(cat, "#666666")
            proc_by_month = monthly.set_index("YearMonth")[proc]
            y_vals = [proc_by_month.get(ym) for ym in months]
            ax.plot(x, y_vals, marker="o", markersize=4, label=proc, color=color, linewidth=1.5, zorder=3)

        ax.plot(
            x,
            total_yields,
            marker="s",
            markersize=5,
            label="总Yield",
            color="#C00000",
            linewidth=2.5,
            linestyle="--",
            zorder=3,
        )
        for xi, y in zip(x, total_yields):
            if y is not None:
                ax.annotate(
                    f"{y * 100:.1f}%",
                    (xi, y),
                    textcoords="offset points",
                    xytext=(0, -12),
                    ha="center",
                    fontsize=7,
                    color="#C00000",
                )

        ax.set_xticks(x)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.set_ylabel("Yield")
        ax.set_xlabel("YearMonth")
        ax.set_title("各工序月度良率及总Yield（总Yield = 常温 × 低温 × 高温）")
        ax.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter(1.0))
        _merge_legends(ax, ax_bar, loc="upper right", fontsize=8, ncol=2)
        ax.grid(True, alpha=0.3, zorder=0)
        ax.margins(y=0.12)
        self.chart_figs["all"].tight_layout()
        self.chart_canvases["all"].draw()

    def _draw_category_chart(self):
        self.chart_figs["cat"].clear()
        ax = self.chart_figs["cat"].add_subplot(111)
        merged = self.merged_df
        months, total_yields = get_monthly_total_yields(merged)
        x = list(range(len(months)))
        cat_test_qty = get_monthly_test_qty_by_category(merged, months)

        ax_bar = ax.twinx()
        bottom = [0] * len(months)
        for cat in TEMP_LABELS:
            ax_bar.bar(
                x,
                cat_test_qty[cat],
                bottom=bottom,
                width=0.55,
                alpha=0.22,
                color=CATEGORY_COLORS[cat],
                label=f"{cat}TestQty",
                zorder=1,
            )
            bottom = [b + q for b, q in zip(bottom, cat_test_qty[cat])]
        ax_bar.set_ylabel("TestQty")
        ax_bar.yaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
        ax.set_zorder(2)
        ax.patch.set_visible(False)

        for cat in TEMP_LABELS:
            yields = []
            for ym in months:
                cat_data = merged[(merged["YearMonth"] == ym) & (merged["温度类型"] == cat)]
                if cat_data.empty or cat_data["TestQty"].sum() == 0:
                    yields.append(None)
                else:
                    yields.append(cat_data["PassQty"].sum() / cat_data["TestQty"].sum())
            ax.plot(
                x,
                yields,
                marker="o",
                markersize=5,
                label=f"{cat}工序",
                color=CATEGORY_COLORS[cat],
                linewidth=2,
                zorder=3,
            )
            for xi, y in zip(x, yields):
                if y is not None:
                    ax.annotate(
                        f"{y * 100:.1f}%",
                        (xi, y),
                        textcoords="offset points",
                        xytext=(0, 8),
                        ha="center",
                        fontsize=7,
                        color=CATEGORY_COLORS[cat],
                    )

        ax.plot(
            x,
            total_yields,
            marker="s",
            markersize=5,
            label="总Yield",
            color="#C00000",
            linewidth=2.5,
            linestyle="--",
            zorder=3,
        )
        for xi, y in zip(x, total_yields):
            if y is not None:
                ax.annotate(
                    f"{y * 100:.1f}%",
                    (xi, y),
                    textcoords="offset points",
                    xytext=(0, -12),
                    ha="center",
                    fontsize=7,
                    color="#C00000",
                )

        ax.set_xticks(x)
        ax.set_xticklabels(months, rotation=45, ha="right")
        ax.set_ylabel("Yield")
        ax.set_xlabel("YearMonth")
        ax.set_title("常温 / 低温 / 高温 工序月度合并良率及总Yield（总Yield = 常温 × 低温 × 高温）")
        ax.yaxis.set_major_formatter(matplotlib.ticker.PercentFormatter(1.0))
        _merge_legends(ax, ax_bar, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0, fontsize=8)
        ax.grid(True, alpha=0.3, zorder=0)
        ax.margins(y=0.12)
        self.chart_figs["cat"].subplots_adjust(right=0.78)
        self.chart_figs["cat"].tight_layout(rect=[0, 0, 0.78, 1])
        self.chart_canvases["cat"].draw()


def main():
    app = YieldAnalyzerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
